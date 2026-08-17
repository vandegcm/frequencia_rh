import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl

from calculadora_horas_extras import CalculadoraHorasExtras
from servicos.gerador_relatorios import GeradorRelatorios
from servicos.otimizador_limites import OtimizadorLimites
from servicos.processador_extras import ProcessadorExtras


class ConfiguracaoBasesTests(unittest.TestCase):
    def test_composicao_padrao_separa_gas_da_emenda_19(self):
        lancamentos = [
            ('1005', 'Salário-base', '3000', ''),
            ('1923', 'GAS', '1000', ''),
            ('1056', 'Emenda 19', '200', ''),
            ('1059', 'Tempo de serviço', '300', ''),
            ('9999', 'Rubrica não configurada', '', ''),
        ]

        sv1, sv2 = ProcessadorExtras.calcular_bases_meta4(lancamentos)

        self.assertEqual(sv1, 4500)
        self.assertEqual(sv2, 3500)

    def test_soma_lancamentos_repetidos_da_mesma_rubrica(self):
        lancamentos = [
            ('1005', 'Salário-base', '3000', ''),
            ('1056', 'Emenda 19', '100', ''),
            ('1056', 'Emenda 19 complementar', '50', ''),
        ]

        sv1, sv2 = ProcessadorExtras.calcular_bases_meta4(
            lancamentos,
            codigos_sv1=('1005', '1056'),
            codigos_sv2=('1005',),
        )

        self.assertEqual(sv1, 3150)
        self.assertEqual(sv2, 3000)


class CalculadoraHorasExtrasTests(unittest.TestCase):
    def test_carga_textual_40_usa_divisor_200(self):
        calc = CalculadoraHorasExtras(carga='40', base_sv1=3000, base_sv2=2500)

        self.assertEqual(calc.ch, 200)

    def test_carga_textual_20_usa_divisor_100(self):
        calc = CalculadoraHorasExtras(carga='20', base_sv1=3000, base_sv2=2500)

        self.assertEqual(calc.ch, 100)

    def test_sobreaviso_fora_do_redutor_e_somado_uma_vez(self):
        calc = CalculadoraHorasExtras(
            carga=40,
            base_sv1=3000,
            base_sv2=3000,
            he_dom_fer=40,
            sobreaviso=30,
            inclui_sobreaviso_no_redutor=False,
        )

        self.assertAlmostEqual(calc.she(), 1200)
        self.assertAlmostEqual(calc.val_sobreaviso(), 149.985)
        self.assertAlmostEqual(calc.base_redutor(), 1200)
        self.assertAlmostEqual(calc.redutor_aplicado(), 200.1)
        self.assertAlmostEqual(calc.total_bruto_geral(), 1349.985)
        self.assertAlmostEqual(calc.total_liquido(), 1149.885)

    def test_valores_unitarios_expostos_no_resultado(self):
        calc = CalculadoraHorasExtras(
            quadro='QPSS', carga=40, base_sv1=3000, base_sv2=2400
        )

        resultado = calc.calcular()

        self.assertEqual(resultado['valor_hora'], 15)
        self.assertAlmostEqual(resultado['valor_hora_sobreaviso'], 3.9996)

    def test_sobreaviso_dentro_do_redutor_sem_duplicacao(self):
        calc = CalculadoraHorasExtras(
            carga=40,
            base_sv1=3000,
            base_sv2=3000,
            he_dom_fer=40,
            sobreaviso=30,
            inclui_sobreaviso_no_redutor=True,
        )

        self.assertAlmostEqual(calc.base_redutor(), 1349.985)
        self.assertAlmostEqual(calc.redutor_aplicado(), 350.085)
        self.assertAlmostEqual(calc.total_bruto_geral(), 1349.985)
        self.assertAlmostEqual(calc.total_liquido(), 999.9)

    def test_otimizador_nao_limita_sobreaviso_quando_fora_do_redutor(self):
        calc = CalculadoraHorasExtras(
            carga=40,
            base_sv1=3000,
            base_sv2=3000,
            inclui_sobreaviso_no_redutor=False,
        )
        horas = {'50d': 0, '50n': 0, '100': 40, 'sa': 30}

        pagar, banco = OtimizadorLimites.calcular_horas_a_pagar(calc, horas)

        self.assertEqual(pagar['100'], 33)
        self.assertEqual(banco['100'], 7)
        self.assertEqual(pagar['sa'], 30)
        self.assertEqual(banco['sa'], 0)

    def test_otimizador_limita_sobreaviso_quando_dentro_do_redutor(self):
        calc = CalculadoraHorasExtras(
            carga=40,
            base_sv1=3000,
            base_sv2=3000,
            inclui_sobreaviso_no_redutor=True,
        )
        horas = {'50d': 0, '50n': 0, '100': 40, 'sa': 30}

        pagar, banco = OtimizadorLimites.calcular_horas_a_pagar(calc, horas)

        self.assertEqual(pagar['100'], 33)
        self.assertEqual(banco['100'], 7)
        self.assertEqual(pagar['sa'], 1)
        self.assertEqual(banco['sa'], 29)


class RelatorioExtrasTests(unittest.TestCase):
    def test_relatorio_mantem_linhas_de_problemas(self):
        with TemporaryDirectory() as pasta:
            GeradorRelatorios.gerar_relatorio_extras(
                pasta_salvar=pasta,
                lista_ids=[111, 222],
                l_bh={'50d': [0, 0], '50n': [0, 0], '100': [0, 0], 'sa': [0, 0]},
                l_hr={'50d': [0, 0], '50n': [0, 0], '100': [0, 0], 'sa': [0, 0]},
                dados_funcionarios_meta4=[],
                calculos={222: {'duplo_vinculo': True, 'nome': 'Servidor Duplicado'}},
                problemas={'dados_indisponiveis': [111], 'duplo_vinculo': [222]},
            )

            arquivo = next(Path(pasta).glob('*.xlsx'))
            aba = openpyxl.load_workbook(arquivo, data_only=True)['HE']

            self.assertEqual((aba['A3'].value, aba['B3'].value), (111, 'DADOS INDISPONIVEIS'))
            self.assertEqual(aba['A4'].value, 222)
            self.assertEqual(aba['B4'].value, 'Servidor Duplicado')
            self.assertEqual(aba['C4'].value, 'DUPLO VÍNCULO')

    def test_relatorio_usa_estrutura_legada_com_quadro(self):
        calc = CalculadoraHorasExtras(
            carga=40,
            base_sv1=3000,
            base_sv2=3000,
            sobreaviso=30,
            inclui_sobreaviso_no_redutor=False,
        )
        resultado = calc.calcular()
        resultado.update({
            'nome': 'Servidor Teste',
            'funcao': 'Teste',
            'quadro': 'QPSS',
            'horas_pagas': {'50d': 0, '50n': 0, '100': 0, 'sa': 30},
            'novo_saldo': {'50d': 0, '50n': 0, '100': 0, 'sa': 0},
        })

        with TemporaryDirectory() as pasta:
            GeradorRelatorios.gerar_relatorio_extras(
                pasta_salvar=pasta,
                lista_ids=[123],
                l_bh={'50d': [0], '50n': [0], '100': [0], 'sa': [0]},
                l_hr={'50d': [0], '50n': [0], '100': [0], 'sa': [30]},
                dados_funcionarios_meta4=[],
                calculos={123: resultado},
                problemas={'dados_indisponiveis': [], 'duplo_vinculo': []},
            )

            arquivo = next(Path(pasta).glob('*.xlsx'))
            aba = openpyxl.load_workbook(arquivo, data_only=True)['HE']

            self.assertEqual(aba.max_column, 33)
            self.assertEqual(aba['D2'].value, 'Quadro')
            self.assertEqual(aba['E2'].value, 'Valor Sobreaviso')
            self.assertEqual(aba['F2'].value, 'Valor Hora')
            self.assertEqual(aba['G2'].value, 'Limite Pagar')
            self.assertEqual(aba['P1'].value, 'Total Horas')
            self.assertEqual(aba['T1'].value, 'Horas a Pagar')
            self.assertEqual(aba['X1'].value, 'Valor Pago')
            self.assertEqual(aba['AB1'].value, 'Total Pago')
            self.assertEqual(aba['AC1'].value, 'Dif.')
            self.assertEqual(aba['AD1'].value, 'Saldo Atualizado')
            self.assertAlmostEqual(aba['E3'].value, 4.9995)
            self.assertEqual(aba['F3'].value, 15)
            self.assertAlmostEqual(aba['G3'].value, 999.9)
            self.assertEqual(aba['S3'].value, 30)
            self.assertEqual(aba['W3'].value, 30)
            self.assertAlmostEqual(aba['AA3'].value, 149.985)
            self.assertAlmostEqual(aba['AB3'].value, 149.985)
            self.assertAlmostEqual(aba['AC3'].value, 849.915)
            self.assertEqual(aba['AG3'].value, 0)

            aba_formulas = openpyxl.load_workbook(arquivo, data_only=False)['HE']
            self.assertEqual(aba_formulas['P3'].value, '=H3+L3')
            self.assertEqual(aba_formulas['S3'].value, '=K3+O3')
            self.assertEqual(aba_formulas['AB3'].value, '=SUM(X3:AA3)')
            self.assertEqual(aba_formulas['AC3'].value, '=G3-AB3')
            self.assertEqual(aba_formulas['AD3'].value, '=P3-T3')
            self.assertEqual(aba_formulas['AG3'].value, '=S3-W3')
            self.assertIn('A1:G1', {str(rng) for rng in aba_formulas.merged_cells.ranges})
            self.assertIn('AD1:AG1', {str(rng) for rng in aba_formulas.merged_cells.ranges})

            self.assertEqual(aba_formulas['P1'].fill.fgColor.rgb, 'FFE4DFEC')
            self.assertEqual(aba_formulas['P2'].fill.fgColor.rgb, 'FFE4DFEC')
            self.assertEqual(aba_formulas['P3'].fill.fgColor.rgb, 'FFE4DFEC')
            self.assertEqual(aba_formulas['X1'].fill.fgColor.rgb, 'FFCCC1D9')
            self.assertEqual(aba_formulas['X2'].fill.fgColor.rgb, 'FFCCC1D9')
            self.assertEqual(aba_formulas['X3'].fill.fgColor.rgb, 'FFCCC1D9')
            self.assertEqual(aba_formulas['AB1'].fill.fgColor.rgb, 'FF92D050')
            self.assertEqual(aba_formulas['AB3'].fill.fgColor.rgb, 'FF92D050')
            self.assertEqual(aba_formulas['AC1'].fill.fgColor.rgb, 'FF92D050')
            self.assertEqual(aba_formulas['AC3'].fill.fgColor.rgb, 'FF92D050')
            self.assertEqual(aba_formulas['AD1'].fill.fgColor.rgb, 'FFFFC000')
            self.assertEqual(aba_formulas['AD2'].fill.fgColor.rgb, 'FFFFC000')
            self.assertEqual(aba_formulas['AD3'].fill.fgColor.rgb, 'FFFFC000')

            for celula in ('E3', 'F3', 'G3', 'X3', 'Y3', 'Z3', 'AA3', 'AB3', 'AC3'):
                self.assertEqual(aba_formulas[celula].number_format, '0.00')


if __name__ == '__main__':
    unittest.main()
