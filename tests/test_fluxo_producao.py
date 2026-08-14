import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

import main
from modelos.funcionario import FuncionarioFrequencia
from servicos.gerador_relatorios import GeradorRelatorios
from servicos.leitor_planilhas import LeitorPlanilhas
from servicos.processador_extras import ProcessadorExtras
from servicos.processador_frequencia import ProcessadorFrequencia
from servicos.resultado import ResultadoProcessamento
from utils.formatadores import validar_competencia


class ValidacaoCompetenciaTests(unittest.TestCase):
    def test_normaliza_mes_e_ano(self):
        self.assertEqual(("01", "2026"), validar_competencia("1", "2026"))

    def test_rejeita_mes_fora_do_intervalo(self):
        with self.assertRaisesRegex(ValueError, "entre 1 e 12"):
            validar_competencia("13", "2026")

    def test_rejeita_ano_sem_quatro_digitos(self):
        with self.assertRaisesRegex(ValueError, "quatro digitos"):
            validar_competencia("8", "26")


class MenuTests(unittest.TestCase):
    @patch("main.mostrar_resultado")
    def test_executar_operacao_repassa_resultado_sem_falso_sucesso(self, mostrar):
        esperado = ResultadoProcessamento(False, "Arquivo ausente")

        main.executar_operacao(lambda: esperado)

        mostrar.assert_called_once_with(esperado)

    @patch("main.mostrar_resultado")
    def test_executar_operacao_converte_erro_esperado_em_falha(self, mostrar):
        def operacao():
            raise PermissionError("Sem permissao")

        main.executar_operacao(operacao)

        resultado = mostrar.call_args.args[0]
        self.assertFalse(resultado.sucesso)
        self.assertIn("Sem permissao", resultado.mensagem)

    @patch("builtins.input", return_value="")
    @patch("builtins.print")
    def test_mostrar_resultado_lista_arquivo_gerado(self, imprimir, _input):
        resultado = ResultadoProcessamento(
            True,
            "Concluido",
            (Path("relatorio.xlsx"),),
        )

        main.mostrar_resultado(resultado)

        textos = " ".join(str(chamada.args[0]) for chamada in imprimir.call_args_list)
        self.assertIn("SUCESSO", textos)
        self.assertIn("relatorio.xlsx", textos)


class ProcessadoresTests(unittest.TestCase):
    def test_lista_informa_arquivo_de_funcionarios_ausente(self):
        with tempfile.TemporaryDirectory() as pasta:
            referencia = Path(pasta) / "SESA"
            referencia.mkdir()

            resultado = ProcessadorFrequencia.processar_lista_frequencia(
                str(referencia), "08", "2026"
            )

        self.assertFalse(resultado.sucesso)
        self.assertIn("funcionarios.xlsx", resultado.mensagem)

    def test_boletim_ausente_nao_cria_estrutura_de_pastas(self):
        with tempfile.TemporaryDirectory() as pasta:
            referencia = Path(pasta) / "FUNEAS"
            referencia.mkdir()

            resultado = ProcessadorFrequencia.processar_boletins(
                str(referencia), "08", "2026"
            )

            self.assertFalse((referencia / "2026").exists())

        self.assertFalse(resultado.sucesso)

    def test_extras_informa_competencia_ausente(self):
        with tempfile.TemporaryDirectory() as pasta:
            raiz = Path(pasta)
            with (
                patch("servicos.processador_extras.PASTA_RELATORIOS_META4", raiz / "meta4"),
                patch("servicos.processador_extras.PASTA_EXTRAS", raiz / "extras"),
            ):
                resultado = ProcessadorExtras.processar_relatorios_meta4_extras(8, 2026)

        self.assertFalse(resultado.sucesso)
        self.assertIn("Pastas da competencia", resultado.mensagem)


class LeitorPlanilhasTests(unittest.TestCase):
    def test_planilha_gerada_pode_ser_lida_sem_deslocar_colunas(self):
        funcionario = FuncionarioFrequencia(
            nome="Pessoa Teste",
            rg="123",
            admissao="01/01/2020",
            funcao="Tecnico",
            setor="SETOR",
            fonte="SESA",
            id_funcionario=1,
        )
        with tempfile.TemporaryDirectory() as pasta:
            (caminho,) = GeradorRelatorios.gerar_lista_frequencia(
                pasta, [funcionario]
            )
            arquivo = openpyxl.load_workbook(caminho)
            aba = arquivo.worksheets[0]
            aba.cell(row=2, column=7).value = "Licenca"
            aba.cell(row=2, column=8).value = "Ferias"
            aba.cell(row=2, column=17).value = "Regular"
            arquivo.save(caminho)
            arquivo.close()

            (lido,) = LeitorPlanilhas.ler_dados_frequencia(str(caminho))

        self.assertEqual("Pessoa Teste", lido.nome)
        self.assertEqual("Licenca", lido.afastamentos_a)
        self.assertEqual("Ferias", lido.afastamentos_b)
        self.assertEqual("Regular", lido.frequencia)

    def test_ignora_arquivos_que_nao_sao_csv(self):
        with tempfile.TemporaryDirectory() as pasta:
            Path(pasta, "ignorar.txt").write_text("conteudo invalido", encoding="utf-8")

            funcionarios = LeitorPlanilhas.ler_relatorios_meta4_csv(pasta)

        self.assertEqual([], funcionarios)

    def test_exige_abas_obrigatorias_na_planilha_de_extras(self):
        with tempfile.TemporaryDirectory() as pasta:
            caminho = Path(pasta) / "extras.xlsx"
            arquivo = openpyxl.Workbook()
            arquivo.save(caminho)
            arquivo.close()

            with self.assertRaisesRegex(KeyError, "GROUND_ZERO"):
                LeitorPlanilhas.ler_banco_horas(str(caminho))


if __name__ == "__main__":
    unittest.main()
