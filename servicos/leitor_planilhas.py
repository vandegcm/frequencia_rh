import openpyxl
import re
import logging
from pathlib import Path
from datetime import datetime, time, timedelta
from typing import List, Tuple, Dict

from modelos.funcionario import FuncionarioFrequencia, FuncionarioMeta4
from config import HOSPITAL_REGIONAL

LOGGER = logging.getLogger(__name__)

class LeitorPlanilhas:
    
    @staticmethod
    def ler_dados_frequencia(caminho_arquivo: str) -> List[FuncionarioFrequencia]:
        arquivo = openpyxl.load_workbook(caminho_arquivo, data_only=True)
        aba = arquivo.worksheets[0]
        fonte_setor = aba.title
        
        if '_' in fonte_setor:
            fonte, setor = fonte_setor.split('_', 1)
        else:
            fonte = fonte_setor
            setor = ""

        funcionarios = []
        # A planilha gerada pela aplicacao possui cabecalho na linha 1.
        for linha in range(2, aba.max_row + 1):
            if not aba.cell(row=linha, column=2).value:
                continue
            func = FuncionarioFrequencia(
                nome=aba.cell(row=linha, column=2).value,
                rg=aba.cell(row=linha, column=3).value,
                admissao=aba.cell(row=linha, column=4).value,
                funcao=aba.cell(row=linha, column=5).value,
                dias_trabalhados=aba.cell(row=linha, column=6).value,
                afastamentos_a=aba.cell(row=linha, column=7).value,
                afastamentos_b=aba.cell(row=linha, column=8).value,
                atraso_saida=aba.cell(row=linha, column=9).value,
                dias_falta=aba.cell(row=linha, column=10).value,
                adicional_noturno=aba.cell(row=linha, column=11).value,
                he_50_d=aba.cell(row=linha, column=12).value,
                he_50_n=aba.cell(row=linha, column=13).value,
                he_100=aba.cell(row=linha, column=14).value,
                sobreaviso=aba.cell(row=linha, column=15).value,
                observacoes=aba.cell(row=linha, column=16).value,
                frequencia=aba.cell(row=linha, column=17).value,
                setor=setor,
                fonte=fonte
            )
            funcionarios.append(func)
        arquivo.close()
        return funcionarios

    @staticmethod
    def buscar_funcionarios_xlsx(caminho_arquivo: str) -> Tuple[List[FuncionarioFrequencia], List[FuncionarioFrequencia]]:
        arquivo = openpyxl.load_workbook(caminho_arquivo, data_only=True)
        aba = arquivo.worksheets[0]
        
        funcionarios_sesa = []
        funcionarios_funeas = []
        
        for linha in range(2, aba.max_row + 1):
            if aba.cell(row=linha, column=9).value == "sim":
                fonte = aba.cell(row=linha, column=8).value
                id_val = aba.cell(row=linha, column=1).value
                func = FuncionarioFrequencia(
                    id_funcionario=int(id_val) if id_val else None,
                    funcao=aba.cell(row=linha, column=2).value,
                    nome=aba.cell(row=linha, column=3).value,
                    rg=aba.cell(row=linha, column=4).value,
                    admissao=aba.cell(row=linha, column=5).value,
                    exoneracao=aba.cell(row=linha, column=6).value,
                    setor=aba.cell(row=linha, column=7).value,
                    fonte=fonte
                )
                
                if fonte == "SESA":
                    funcionarios_sesa.append(func)
                elif fonte == "FUNEAS":
                    funcionarios_funeas.append(func)

        arquivo.close()
        return funcionarios_sesa, funcionarios_funeas
        
    @staticmethod
    def ler_relatorios_meta4_csv(pasta_relatorios: str) -> List[FuncionarioMeta4]:
        padrao_id = re.compile(r'\d+')
        funcionarios_dict = {}
        
        arquivos = sorted(Path(pasta_relatorios).iterdir())

        for caminho_relatorio in arquivos:
            if not caminho_relatorio.is_file() or caminho_relatorio.suffix.lower() != '.csv':
                continue
            caminho_completo = str(caminho_relatorio)
            try:
                with open(caminho_completo, 'r', encoding='utf-8') as relatorio_aberto:
                    linhas = relatorio_aberto.readlines()
            except UnicodeDecodeError:
                with open(caminho_completo, 'r', encoding='latin-1') as relatorio_aberto:
                    linhas = relatorio_aberto.readlines()

            for linha in linhas:
                linha_split = linha.replace('\n', '').split(',')
                if len(linha_split) == 22 and linha_split[16] == HOSPITAL_REGIONAL:
                    n_id = int(linha_split[2])
                    
                    if n_id not in funcionarios_dict:
                        func = FuncionarioMeta4(
                            nome=linha_split[0], rg=linha_split[1], n_id=n_id,
                            t_emp=linha_split[3], nascimento=linha_split[4], pis=linha_split[5],
                            cpf=linha_split[6], admissao=linha_split[7], id_ato_formal=linha_split[8],
                            carga_horaria=linha_split[9], cargo=linha_split[10], funcao=linha_split[11],
                            classe=linha_split[12], referencia_classe=linha_split[13],
                            cidade=linha_split[14], estado=linha_split[15], quadro=linha_split[17]
                        )
                        funcionarios_dict[n_id] = func
                        
                    descricao = linha_split[18]
                    match = re.search(padrao_id, descricao)
                    if match:
                        id_descricao = match.group()
                        descricao = descricao.replace(id_descricao, '').strip()
                        vantagem = linha_split[19]
                        desconto = linha_split[20]
                        funcionarios_dict[n_id].lancamentos.append((id_descricao, descricao, vantagem, desconto))
                        
        return list(funcionarios_dict.values())
        
    @staticmethod
    def ler_banco_horas(caminho_arquivo: str) -> Dict[int, tuple]:
        arquivo = openpyxl.load_workbook(caminho_arquivo, data_only=True)
        if 'GROUND_ZERO' not in arquivo.sheetnames:
            arquivo.close()
            raise KeyError("A planilha deve conter a aba GROUND_ZERO.")
        aba = arquivo['GROUND_ZERO']
        
        banco = {}
        for linha in range(2, aba.max_row + 1):
            id_val = aba.cell(row=linha, column=2).value
            if not id_val:
                continue
            id_funcionario = int(id_val)
            h50d = aba.cell(row=linha, column=5).value or 0.0
            h50n = aba.cell(row=linha, column=6).value or 0.0
            h100 = aba.cell(row=linha, column=7).value or 0.0
            hsa = aba.cell(row=linha, column=8).value or 0.0
            banco[id_funcionario] = (float(h50d), float(h50n), float(h100), float(hsa))

        arquivo.close()
        return banco

    @staticmethod
    def ler_plantoes(caminho_arquivo: str, nome_aba: str) -> Dict[int, List[Tuple[datetime, datetime]]]:
        arquivo = openpyxl.load_workbook(caminho_arquivo, data_only=True)
        if nome_aba not in arquivo.sheetnames:
            arquivo.close()
            raise KeyError(f"A planilha deve conter a aba {nome_aba}.")
            
        aba = arquivo[nome_aba]
        plantoes = {}
        
        for linha in range(2, aba.max_row + 1):
            id_val = aba.cell(row=linha, column=1).value
            if not id_val:
                continue
                
            id_funcionario = int(id_val)
            dia = aba.cell(row=linha, column=3).value
            inicio = aba.cell(row=linha, column=6).value
            horas_autorizadas = aba.cell(row=linha, column=9).value
            
            if not dia or not inicio or not horas_autorizadas:
                continue
                
            try:
                # Tratar dia
                if isinstance(dia, datetime):
                    data_dia = dia
                else:
                    # Fallback (caso venha string ou outro formato, ignoramos neste protótipo)
                    continue
                    
                # Tratar inicio
                if isinstance(inicio, time):
                    hi_time = inicio
                elif isinstance(inicio, datetime):
                    hi_time = inicio.time()
                elif isinstance(inicio, str):
                    hi, mi = inicio.split(':')
                    hi_time = time(hour=int(hi), minute=int(mi))
                else:
                    continue
                    
                # Tratar horas_autorizadas
                if isinstance(horas_autorizadas, time):
                    ha_time = horas_autorizadas
                elif isinstance(horas_autorizadas, datetime):
                    ha_time = horas_autorizadas.time()
                elif isinstance(horas_autorizadas, str):
                    hai, mai = horas_autorizadas.split(':')
                    ha_time = time(hour=int(hai), minute=int(mai))
                else:
                    continue
                    
                horario_entrada = datetime.combine(data_dia.date(), hi_time)
                
                # Somar horas_autorizadas
                delta_autorizado = timedelta(hours=ha_time.hour, minutes=ha_time.minute, seconds=ha_time.second)
                saida_autorizada = horario_entrada + delta_autorizado
                
                if id_funcionario not in plantoes:
                    plantoes[id_funcionario] = []
                plantoes[id_funcionario].append((horario_entrada, saida_autorizada))
            except (AttributeError, TypeError, ValueError) as erro:
                LOGGER.warning(
                    "Linha %s ignorada na aba %s: %s", linha, nome_aba, erro
                )

        arquivo.close()
        return plantoes

    @staticmethod
    def ler_feriados(caminho_arquivo: str) -> List[datetime]:
        arquivo = openpyxl.load_workbook(caminho_arquivo, data_only=True)
        if 'FERIADOS' not in arquivo.sheetnames:
            arquivo.close()
            raise KeyError("A planilha deve conter a aba FERIADOS.")
        aba = arquivo['FERIADOS']
        feriados = []
        
        for linha in range(2, aba.max_row + 1):
            feriado = aba.cell(row=linha, column=1).value
            if feriado and isinstance(feriado, datetime):
                feriados.append(feriado)

        arquivo.close()
        return feriados
